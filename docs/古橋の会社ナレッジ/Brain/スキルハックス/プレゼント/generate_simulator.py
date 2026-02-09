#!/usr/bin/env python3
"""副業収支シミュレーター スプレッドシート生成スクリプト"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Color palette ──
DARK_NAVY = "1A1A2E"
GOLD = "D4A853"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MID_GRAY = "E0E0E0"
DARK_TEXT = "333333"
RED = "E74C3C"
GREEN = "27AE60"
BLUE = "3498DB"

# ── Reusable styles ──
title_font = Font(name="Arial", size=16, bold=True, color=DARK_NAVY)
header_font = Font(name="Arial", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
gold_font = Font(name="Arial", size=12, bold=True, color=DARK_NAVY)
input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
input_font = Font(name="Arial", size=12, color=DARK_NAVY)
label_font = Font(name="Arial", size=11, color=DARK_TEXT)
result_font = Font(name="Arial", size=14, bold=True, color=DARK_NAVY)
note_font = Font(name="Arial", size=9, italic=True, color="888888")
section_font = Font(name="Arial", size=13, bold=True, color=GOLD)
thin_border = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)

def style_range(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if number_format: cell.number_format = number_format
    return cell


# ============================================================
# Sheet 1: 収支シミュレーター（メイン）
# ============================================================
ws1 = wb.active
ws1.title = "収支シミュレーター"
ws1.sheet_properties.tabColor = DARK_NAVY

# Column widths
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 30
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 20
ws1.column_dimensions["E"].width = 20
ws1.column_dimensions["F"].width = 3
ws1.column_dimensions["G"].width = 25
ws1.column_dimensions["H"].width = 20

# ── Title ──
ws1.merge_cells("B2:E2")
style_range(ws1, 2, 2, "💰 副業収支シミュレーター", title_font, alignment=Alignment(horizontal="center"))
ws1.merge_cells("B3:E3")
style_range(ws1, 3, 2, "あなたの副業収入をシミュレーションしましょう", 
            Font(name="Arial", size=10, italic=True, color="888888"),
            alignment=Alignment(horizontal="center"))

# ── Section 1: 基本情報入力 ──
style_range(ws1, 5, 2, "📋 STEP 1：基本情報を入力", section_font)

labels_1 = [
    ("スキル種別", "動画編集", "動画編集 / デザイン / プログラミング / ライティング から選択"),
    ("1案件あたりの単価（円）", 5000, "例）動画編集5,000円、LP制作50,000円"),
    ("月の作業可能時間（時間）", 40, "例）平日2h×20日 = 40時間"),
    ("1案件あたりの作業時間（時間）", 3, "例）10分動画の編集 = 約3時間"),
    ("月間経費（円）", 2000, "Adobe月額、通信費 等"),
]

for i, (label, default, note) in enumerate(labels_1):
    row = 7 + i
    style_range(ws1, row, 2, label, label_font, border=thin_border)
    style_range(ws1, row, 3, default, input_font, input_fill, border=thin_border,
                number_format="#,##0" if isinstance(default, int) else None)
    style_range(ws1, row, 4, note, note_font)

# ── Section 2: 自動計算結果 ──
style_range(ws1, 14, 2, "📊 STEP 2：シミュレーション結果", section_font)

# Formulas
results = [
    ("月間案件数（自動計算）", "=ROUNDDOWN(C9/C10,0)", "件", "作業可能時間 ÷ 1案件の作業時間"),
    ("月間売上（税込）", "=C8*C16", "円", "単価 × 月間案件数"),
    ("月間経費", "=C11", "円", None),
    ("月間利益（手取り目安）", "=C17-C18", "円", "売上 − 経費"),
    ("年間利益（概算）", "=C19*12", "円", None),
    ("実質時給", "=ROUND(C19/C9,0)", "円/時間", "月間利益 ÷ 月の作業時間"),
]

for i, (label, formula, unit, note) in enumerate(results):
    row = 16 + i
    style_range(ws1, row, 2, label, label_font, border=thin_border)
    c = style_range(ws1, row, 3, None, result_font, border=thin_border)
    c.value = formula
    c.number_format = "#,##0"
    style_range(ws1, row, 4, unit, label_font)
    if note:
        style_range(ws1, row, 5, note, note_font)

# Highlight profit row
for col in range(2, 5):
    ws1.cell(row=19, column=col).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ws1.cell(row=19, column=3).font = Font(name="Arial", size=14, bold=True, color=GREEN)

# ── Section 3: キャッシュフローの4つの箱 ──
style_range(ws1, 24, 2, "🏦 STEP 3：キャッシュフローの4つの箱", section_font)

cf_labels = [
    (26, "収入", "=C19", GREEN, "副業の月間利益"),
    (27, "支出（生活費を想定）", 150000, RED, "家賃・食費・固定費 等（入力してください）"),
    (28, "貯蓄に回せる額", "=C26-C27", BLUE, "収入 − 支出"),
    (29, "資産運用に回す額", "=C28*0.3", GOLD, "貯蓄の30%を投資に回す想定"),
]

for row, label, value, color, note in cf_labels:
    style_range(ws1, row, 2, label, label_font, border=thin_border)
    c = ws1.cell(row=row, column=3)
    c.value = value
    c.number_format = "#,##0"
    c.border = thin_border
    if isinstance(value, str) and value.startswith("="):
        c.font = Font(name="Arial", size=12, bold=True, color=color)
    else:
        c.font = input_font
        c.fill = input_fill
    style_range(ws1, row, 4, "円", label_font)
    style_range(ws1, row, 5, note, note_font)

# ── Section 4: 成長シミュレーション ──
style_range(ws1, 32, 2, "📈 STEP 4：成長シミュレーション（12ヶ月予測）", section_font)

# Header row
growth_headers = ["月", "単価（円）", "月間案件数", "月間売上", "経費", "月間利益", "累計利益"]
for j, h in enumerate(growth_headers):
    style_range(ws1, 34, 2+j, h, header_font, header_fill, Alignment(horizontal="center"), thin_border)

# Month 1 (base)
style_range(ws1, 35, 2, "1ヶ月目", label_font, border=thin_border)
ws1.cell(row=35, column=3, value="=C8").number_format = "#,##0"
ws1.cell(row=35, column=4, value="=C16").number_format = "#,##0"
ws1.cell(row=35, column=5, value="=C3*D35").number_format = "#,##0"  # use formula
ws1.cell(row=35, column=5).value = "=C35*D35"
ws1.cell(row=35, column=6, value="=C11").number_format = "#,##0"
ws1.cell(row=35, column=7, value="=E35-F35").number_format = "#,##0"
ws1.cell(row=35, column=8, value="=G35").number_format = "#,##0"

for col in range(3, 9):
    ws1.cell(row=35, column=col).border = thin_border
    ws1.cell(row=35, column=col).font = label_font

# Months 2-12 with growth assumptions
for i in range(1, 12):
    row = 35 + i
    month = i + 1
    style_range(ws1, row, 2, f"{month}ヶ月目", label_font, border=thin_border)
    
    # Price increases: +10% at month 4, +20% at month 7, +30% at month 10
    if month <= 3:
        price_formula = f"=C35"
    elif month <= 6:
        price_formula = f"=ROUND(C35*1.1,0)"
    elif month <= 9:
        price_formula = f"=ROUND(C35*1.3,0)"
    else:
        price_formula = f"=ROUND(C35*1.5,0)"
    
    # Cases increase: +1 at month 3, +2 at month 6, +3 at month 9
    if month <= 3:
        cases_formula = f"=C16"
    elif month <= 6:
        cases_formula = f"=C16+2"
    elif month <= 9:
        cases_formula = f"=C16+4"
    else:
        cases_formula = f"=C16+6"
    
    ws1.cell(row=row, column=3, value=price_formula).number_format = "#,##0"
    ws1.cell(row=row, column=4, value=cases_formula).number_format = "#,##0"
    ws1.cell(row=row, column=5, value=f"=C{row}*D{row}").number_format = "#,##0"
    ws1.cell(row=row, column=6, value=f"=C11").number_format = "#,##0"
    ws1.cell(row=row, column=7, value=f"=E{row}-F{row}").number_format = "#,##0"
    ws1.cell(row=row, column=8, value=f"=H{row-1}+G{row}").number_format = "#,##0"
    
    for col in range(3, 9):
        ws1.cell(row=row, column=col).border = thin_border
        ws1.cell(row=row, column=col).font = label_font

# Growth chart
chart = BarChart()
chart.type = "col"
chart.title = "月間利益の推移（12ヶ月）"
chart.y_axis.title = "円"
chart.x_axis.title = "月"
chart.style = 10

data = Reference(ws1, min_col=7, min_row=34, max_row=46)
cats = Reference(ws1, min_col=2, min_row=35, max_row=46)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 25
chart.height = 12

ws1.add_chart(chart, "B49")

# ── Notes ──
style_range(ws1, 63, 2, "💡 使い方", section_font)
notes = [
    "1. 黄色いセル（STEP 1）に自分の情報を入力してください",
    "2. STEP 2以降は自動計算されます",
    "3. 成長シミュレーションは、3ヶ月ごとに単価10〜20%UP、案件数+2件を想定しています",
    "4. 実際の成長速度は人それぞれです。目安としてご活用ください",
    "5. このシートをGoogleスプレッドシートにアップロードして使えます",
]
for i, note in enumerate(notes):
    style_range(ws1, 65 + i, 2, note, note_font)


# ============================================================
# Sheet 2: スキル別の相場表
# ============================================================
ws2 = wb.create_sheet("スキル別相場表")
ws2.sheet_properties.tabColor = GOLD

ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 20
ws2.column_dimensions["C"].width = 25
ws2.column_dimensions["D"].width = 18
ws2.column_dimensions["E"].width = 18
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 15

ws2.merge_cells("B2:G2")
style_range(ws2, 2, 2, "📊 スキル別 案件相場一覧", title_font, alignment=Alignment(horizontal="center"))

# Headers
market_headers = ["スキル", "案件の種類", "初心者相場", "実績3件後", "実績10件後", "作業時間目安"]
for j, h in enumerate(market_headers):
    style_range(ws2, 4, 2+j, h, header_font, header_fill, Alignment(horizontal="center"), thin_border)

market_data = [
    ["動画編集", "YouTube動画（10分）", "3,000〜5,000円", "5,000〜8,000円", "8,000〜15,000円", "2〜4時間"],
    ["動画編集", "ショート動画", "1,000〜3,000円", "3,000〜5,000円", "5,000〜10,000円", "1〜2時間"],
    ["動画編集", "YouTube運営代行（月額）", "−", "3〜5万円", "5〜10万円", "月20〜40時間"],
    ["デザイン", "バナー1枚", "3,000〜5,000円", "5,000〜10,000円", "10,000〜30,000円", "1〜3時間"],
    ["デザイン", "サムネイル1枚", "1,000〜3,000円", "3,000〜5,000円", "5,000〜10,000円", "30分〜1時間"],
    ["デザイン", "LP デザイン", "3〜5万円", "5〜10万円", "10〜30万円", "10〜20時間"],
    ["プログラミング", "LP コーディング", "3〜5万円", "5〜10万円", "10〜30万円", "5〜15時間"],
    ["プログラミング", "サイト修正", "5,000〜10,000円", "10,000〜30,000円", "30,000〜50,000円", "1〜5時間"],
    ["プログラミング", "WordPress構築", "5〜10万円", "10〜20万円", "20〜50万円", "20〜40時間"],
    ["ライティング", "SEO記事（3,000字）", "3,000円（文字単価1円）", "6,000〜9,000円", "9,000〜15,000円", "3〜5時間"],
    ["ライティング", "SEO記事（5,000字）", "5,000円（文字単価1円）", "10,000〜15,000円", "15,000〜25,000円", "5〜8時間"],
    ["ライティング", "取材記事", "5,000〜10,000円", "10,000〜30,000円", "30,000〜50,000円", "5〜10時間"],
]

for i, row_data in enumerate(market_data):
    row = 5 + i
    for j, val in enumerate(row_data):
        style_range(ws2, row, 2+j, val, label_font, border=thin_border)
    # Alternate row color
    if i % 2 == 0:
        for j in range(len(row_data)):
            ws2.cell(row=row, column=2+j).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")


# ============================================================
# Sheet 3: 目標逆算シート
# ============================================================
ws3 = wb.create_sheet("目標逆算シート")
ws3.sheet_properties.tabColor = GREEN

ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 30
ws3.column_dimensions["E"].width = 20

ws3.merge_cells("B2:E2")
style_range(ws3, 2, 2, "🎯 目標逆算シート", title_font, alignment=Alignment(horizontal="center"))
ws3.merge_cells("B3:E3")
style_range(ws3, 3, 2, "目標月収から、必要な案件数・作業時間を逆算します",
            Font(name="Arial", size=10, italic=True, color="888888"),
            alignment=Alignment(horizontal="center"))

# Input
style_range(ws3, 5, 2, "📋 目標と条件を入力", section_font)

goal_labels = [
    ("目標月収（円）", 100000, "例）100,000円 = 月10万円"),
    ("1案件あたりの単価（円）", 5000, "メインシートの値を参照してもOK"),
    ("1案件あたりの作業時間（時間）", 3, None),
    ("月間経費（円）", 2000, None),
]

for i, (label, default, note) in enumerate(goal_labels):
    row = 7 + i
    style_range(ws3, row, 2, label, label_font, border=thin_border)
    style_range(ws3, row, 3, default, input_font, input_fill, border=thin_border, number_format="#,##0")
    if note:
        style_range(ws3, row, 4, note, note_font)

# Calculation results
style_range(ws3, 13, 2, "📊 逆算結果", section_font)

goal_results = [
    ("必要な月間売上（経費込み）", "=C7+C10", "目標月収 + 経費"),
    ("必要な月間案件数", "=ROUNDUP(C15/C8,0)", "売上 ÷ 単価（切り上げ）"),
    ("必要な月間作業時間", "=C16*C9", "案件数 × 1案件の作業時間"),
    ("1日あたりの作業時間（30日計算）", "=ROUND(C17/30,1)", None),
    ("1日あたりの作業時間（平日20日計算）", "=ROUND(C17/20,1)", None),
    ("実質時給", "=ROUND(C7/C17,0)", "目標月収 ÷ 作業時間"),
]

for i, (label, formula, note) in enumerate(goal_results):
    row = 15 + i
    style_range(ws3, row, 2, label, label_font, border=thin_border)
    c = ws3.cell(row=row, column=3)
    c.value = formula
    c.number_format = "#,##0" if "時間" not in label else "0.0"
    c.font = result_font
    c.border = thin_border
    if note:
        style_range(ws3, row, 4, note, note_font)

# Highlight key row
for col in [2, 3]:
    ws3.cell(row=19, column=col).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

# Judgment
style_range(ws3, 23, 2, "💡 判定", section_font)
style_range(ws3, 25, 2, "1日の作業時間が…", label_font)
judgments = [
    ("1時間以内", "→ 🟢 余裕あり。本業と無理なく両立できます"),
    ("1〜2時間", "→ 🟡 現実的。朝 or 夜に時間を確保すればOK"),
    ("2〜3時間", "→ 🟠 やや多め。休日にまとめて作業する工夫を"),
    ("3時間以上", "→ 🔴 単価UPが必要。スキルの掛け合わせを検討"),
]
for i, (time, judgment) in enumerate(judgments):
    row = 26 + i
    style_range(ws3, row, 2, time, label_font, border=thin_border)
    style_range(ws3, row, 3, judgment, label_font, border=thin_border)
    ws3.merge_cells(f"C{row}:E{row}")


# ============================================================
# Save
# ============================================================
output_path = "/Users/hiroshi/cursor/docs/古橋の会社ナレッジ/Brain/スキルハックス/プレゼント/副業収支シミュレーター.xlsx"
wb.save(output_path)
print(f"✅ スプレッドシートを生成しました: {output_path}")
